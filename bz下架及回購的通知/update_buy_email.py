import os
import openpyxl

# 读取buyEmail.xls文件（先转换为xlsx格式）
try:
    import xlrd
    wb = xlrd.open_workbook('buyEmail.xls')
    sheet = wb.sheet_by_index(0)
    
    # 创建一个新的xlsx文件
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new.active
    
    # 复制原文件内容到新文件
    for row_idx in range(sheet.nrows):
        for col_idx in range(sheet.ncols):
            sheet_new.cell(row=row_idx+1, column=col_idx+1).value = sheet.cell(row_idx, col_idx).value
    
    # 遍历html2目录下的所有HTML文件
    html_dir = 'html2'
    for filename in os.listdir(html_dir):
        if filename.endswith('.html'):
            # 提取文件名中的语言代码（如zh-TW.html -> zh-TW）
            file_lang = filename[:-5]  # 去掉.html后缀
            
            # 查找对应的行
            found = False
            for row_idx in range(1, sheet_new.max_row+1):
                # 检查template_code是否为mail.spot.symbol.delisting且lang匹配
                if sheet_new.cell(row=row_idx, column=1).value == 'mail.spot.symbol.delisting' and sheet_new.cell(row=row_idx, column=2).value == file_lang:
                    # 读取HTML文件内容
                    file_path = os.path.join(html_dir, filename)
                    with open(file_path, 'r', encoding='utf-8') as f:
                        html_content = f.read()
                    
                    # 更新Excel文件中的content列
                    sheet_new.cell(row=row_idx, column=4).value = html_content
                    print(f'已更新: {file_lang}行')
                    found = True
                    break
            
            if not found:
                print(f'未找到对应的行: {file_lang}')
    
    # 保存更新后的Excel文件
    wb_new.save('buyEmail_updated.xlsx')
    print('所有文件处理完成！已保存为buyEmail_updated.xlsx')
    
except ImportError:
    print('xlrd库未安装，请先安装xlrd库：pip install xlrd')
except Exception as e:
    print(f'处理文件时出错: {e}')