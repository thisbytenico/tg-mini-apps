import openpyxl

# 读取buyEmail_updated.xlsx文件
wb_email = openpyxl.load_workbook('buyEmail_updated.xlsx')
sheet_email = wb_email.active

# 读取buy.xlsx文件
wb_buy = openpyxl.load_workbook('buy.xlsx')
sheet_buy = wb_buy.active

# 获取buy.xlsx第一行的语言标识（列标题）
lang_codes = [cell.value for cell in sheet_buy[1]]

# 获取buy.xlsx第五行的内容
row5 = [cell.value for cell in sheet_buy[5]]

# 创建语言到内容的映射
translation_map = {}
for lang, content in zip(lang_codes, row5):
    if lang and content:
        translation_map[lang] = content

# 遍历buyEmail_updated.xlsx文件的所有行
for row_idx in range(1, sheet_email.max_row+1):
    # 检查template_code是否为push.spot.symbol.delisting
    if sheet_email.cell(row=row_idx, column=1).value == 'push.spot.symbol.delisting':
        # 获取lang列的值
        lang = sheet_email.cell(row=row_idx, column=2).value
        
        # 查找对应的翻译内容
        if lang in translation_map:
            target_content = translation_map[lang]
            
            # 更新content列
            sheet_email.cell(row=row_idx, column=4).value = target_content
            print(f'已更新: {lang}行')
        else:
            print(f'未找到对应的翻译内容: {lang}')

# 保存更新后的Excel文件
wb_email.save('buyEmail_updated_final.xlsx')
print('所有文件处理完成！已保存为buyEmail_updated_final.xlsx')